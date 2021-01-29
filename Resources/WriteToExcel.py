from openpyxl import load_workbook


def write_value_to_cell(workbook, sheet, rownum:int ,colnum:int , value1:str):
        wb = load_workbook(filename = workbook)
        sheet_ranges = wb[sheet]
        sheet_ranges.cell(row=rownum+1, column=colnum+1).value = value1
        wb.save(workbook)
